import os
import csv
import time
import datetime
from openpyxl import load_workbook


def formatString(frase):
    frase_general = frase.upper().replace(
          "  ", " ").replace("LUEGO", "LUENGO").replace(
          "_", "").replace("-", "").replace(",", "").replace(".", "").replace("/", "").replace(
          "@", "").replace("$", "").replace("!", "").replace("?", "").replace("¿", "").replace(
          "1", "").replace("2", "").replace("3", "").replace("4", "").replace("5", "").replace(
          "Ñ", "N").replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    return frase_general


def formatDate(date):
    fecha = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    anio = fecha.year
    mes = fecha.month
    dia = fecha.day
    fecha_formateada = datetime.datetime(anio, mes, dia)
    return str(fecha_formateada)


def exeptNombres(nombre):
    lista_excluidos = [
        'Escuela de Contadores Auditores de Santiago'
        # 'CAROLINA ELIZABETH LEMUNGUIR SEPULVEDA',
        # 'CONSUELO VARGAS FIGUEROA'
    ]
    if nombre in lista_excluidos:
        return True
    return False


# Recorrer todos los directorios y archivos de la carpeta actual
def recorrer_archivos():

    # Definir variables o contadores globales
    start_time = time.time()
    cursosSinExito = []
    alumnosSinExito = []
    nombresSinExito = []
    contExito = 0
    contLeidos = 0

    try:
        filepath = "./app/asistencia.xlsx"
        SHEET = "Hoja1"
        wb = load_workbook(filepath)
        sheet = wb[SHEET]
        column_count = sheet.max_column
        row_count = sheet.max_row
        for dirpath, dirnames, filenames in os.walk("./app"):
            if not(dirnames) and (filenames):
                if dirpath == './app/__pycache__':
                    continue
                print("Ruta actual:", dirpath)
                print("")

                # Definir variables o contadores para cada directorio
                columna_fecha_archivo = 0
                for filename in filenames:
                    if(filename == '.DS_Store'):
                        continue
                    # Definir variables o contadores para cada archivo
                    alumnos_marcados = 0
                    alumnos_encontrados = 0

                    # Obtener nombre de la materia
                    list_filename_split = filename.split()
                    del list_filename_split[-4:]
                    asignatura_concat = ' '.join(x for x in list_filename_split)
                    asignatura = formatString(asignatura_concat.lstrip(' '))
                    print(asignatura)

                    # Abrir cada archivo
                    with open(dirpath+'/'+filename) as csv_file:

                        # Delimitar el contenido por ";"
                        csv_reader = csv.reader(csv_file, delimiter=';')
                        line_count = 0
                        for columna in csv_reader:
                            try:

                                # Omitir la primera fila
                                if(line_count == 0):
                                    line_count += 1
                                    continue

                                # Excepciones
                                if exeptNombres(str(columna[0])):
                                    continue

                                # Guardar fecha
                                if(line_count == 1):
                                    fecha_asistencia = formatDate(columna[2])
                                    print(fecha_asistencia)

                                # Guardar cada alumno
                                nombre_alumno = formatString(str(columna[0]))
                                for i in range(3, row_count):
                                    try:
                                        listaNombre = [
                                            str(sheet.cell(row=i, column=6).value),
                                            str(sheet.cell(row=i, column=4).value),
                                            str(sheet.cell(row=i, column=5).value)]
                                        nombre_armado_sin_normalizar = " ".join(listaNombre)
                                        nombre_armado = formatString(nombre_armado_sin_normalizar)
                                        if(nombre_armado == nombre_alumno):
                                            asignatura_armada = formatString(sheet.cell(row=i, column=8).value)
                                            if(asignatura_armada == asignatura):
                                                # Asignar fecha
                                                if columna_fecha_archivo == 0:
                                                    for j in range(9, column_count+1):
                                                        if(str(sheet.cell(row=2, column=j).value) == fecha_asistencia):
                                                            columna_fecha_archivo = j
                                                            sheet.cell(row=i, column=columna_fecha_archivo).value = 1
                                                            alumnos_marcados += 1
                                                            break
                                                else:
                                                    sheet.cell(row=i, column=columna_fecha_archivo).value = 1
                                                    alumnos_marcados += 1
                                                    break
                                    except Exception as e:
                                        print("Error fila: "+str(e))
                                        continue
                                else:
                                    lista = [fecha_asistencia.split()[0], asignatura, nombre_alumno]
                                    alumnosSinExito.append(';'.join(lista))
                                    nombresSinExito.append('.'+nombre_alumno+'.')
                                line_count += 1
                            except Exception as e:
                                line_count += 1
                                print("Error columna: "+str(e))
                                continue
                        alumnos_encontrados += line_count
                    contExito += alumnos_marcados
                    contLeidos += (alumnos_encontrados - 2)
                    print(f'Alumnos encontrados: {alumnos_encontrados - 2}, Alumnos marcados: {alumnos_marcados}')
                    print('')
                    if (alumnos_marcados == 0):
                        cursosSinExito.append(asignatura+" ("+alumnos_encontrados+" asistidos ) "+fecha_asistencia)

        wb.save(filename='actualizado.xlsx')
        print("Se guardo el archivo con el nombre: actualizado.xlsx")
        print(f'{contExito} alumnos agregados de {contLeidos} alumnos totales')
        porcentaje_acertado = round(((contExito/contLeidos)*100), 2)
        tiempo_ejecucion_s = round((time.time() - start_time), 2)
        min_ejecucion = int(tiempo_ejecucion_s / 60)
        seg_ejecucion = int(tiempo_ejecucion_s % 60)
        print(f'Equivalente a : {porcentaje_acertado} % de exito')
        print(f'---Se demoro: {min_ejecucion}:{seg_ejecucion} ---')

        if (cursosSinExito):
            print("Cursos vacios: ")
            print(cursosSinExito)

        if (alumnosSinExito):
            print("")
            print(str(len(alumnosSinExito)) + " alumnos no encontrados")
            for x in alumnosSinExito:
                print(x)
            print("")
            print("nombresSinExito:")
            for y in [ele for ind, ele in enumerate(nombresSinExito, 1) if ele not in nombresSinExito[ind:]]:
                print("{} {}".format(y, nombresSinExito.count(y)))

    except Exception as err:
        print("Error Global: "+str(err))
